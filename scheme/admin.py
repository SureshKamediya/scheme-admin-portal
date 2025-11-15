
from django.contrib import admin
from .models import Scheme, SchemeFiles, Application
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from .models import SchemeFiles, Scheme

from import_export.admin import ImportExportModelAdmin

from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
from .tests import SchemeFactory

# scheme = SchemeFactory.create()
# print(scheme.next_application_number)
# valid_application_data = {
#             'scheme': scheme,
#             'mobile_number': f'9876543210',
#             'applicant_name': 'John Doe',
#             'father_or_husband_name': 'Richard Doe',
#             'dob': date(1990, 1, 1),
#             'id_type': 'AADHAR',
#             'id_number': '123456789012',
#             'pan_number': 'ABCDE1234F',
#             'permanent_address': '123 Main St, City',
#             'permanent_address_pincode': '123456',
#             'postal_address': '123 Main St, City',
#             'postal_address_pincode': '123456',
#             'email': 'john@example.com',
#             'annual_income': 'UP_TO_3L',
#             'payment_mode': 'UPI',
#             'refund_bank_branch_address': 'Branch Address',
#             'dd_date_or_transaction_date' : timezone.now(),
#             'dd_amount' : Decimal(21000.00)
#         }

# application1 = Application.objects.create(**valid_application_data)
# print(scheme.next_application_number, application1)

# scheme = SchemeFactory.create()
# valid_application_data['scheme'] = scheme
# application2 = Application.objects.create(**valid_application_data)
# print(application2)


class SchemeFilesResource(resources.ModelResource):

    Scheme = fields.Field(
        column_name='scheme',
        attribute='Scheme',
        widget=ForeignKeyWidget(Scheme, 'name')
    )

    class Meta:
        model = SchemeFiles
        fields = (
            'id',
            'Scheme',
            'name',
            'file_choice',
            'file',
        )
        export_order = fields



@admin.register(SchemeFiles)
class SchemeFilesAdmin(ImportExportModelAdmin):
    resource_class = SchemeFilesResource
    list_display = ('name', 'Scheme', 'file_choice', 'file')
    list_filter = ('Scheme', 'file_choice', )
    search_fields = ('name', 'Scheme__name')
    # readonly_fields = ('name',)
    
    fieldsets = (
        (None, {
            'fields': ('Scheme', 'file_choice', 'name', 'file')
        }),
    )


class SchemeFilesInline(admin.TabularInline):
    """Inline admin for managing scheme files within the Scheme admin"""
    model = SchemeFiles
    extra = 1
    fields = ('file_choice', 'name', 'file')
    can_delete=True
    # readonly_fields = ('name',)  # Auto-populated based on file_choice


@admin.register(Scheme)
class SchemeAdmin(admin.ModelAdmin):
    list_display = ('name', 'company', 'get_status', 'ews_plot_count', 'Lig_plot_count', 'created_at', 
        'next_application_number',)
    list_filter = ('company', 'created_at', 'application_open_date')
    search_fields = ('name', 'address', 'phone')
    readonly_fields = ('created_at',)
    
    # Add inline for files
    inlines = [SchemeFilesInline]
    
    # Organize fields into fieldsets
    fieldsets = (
        ('Basic Information', {
            'fields': ('company', 'name', 'address', 'phone')
        }),
        ('Plot Configuration', {
            'fields': ('ews_plot_count', 'Lig_plot_count', 'reserved_rate'),
            'description': 'Configure the number of plots and reserved rate'
        }),
        ('Application Settings', {
            'fields': ('application_number_start',),
            'description': """Enter the starting application number for this scheme.
All applications submitted under this scheme will receive sequential numbers beginning from this value.
Please choose a number that leaves enough room for growth — ideally maintain a gap of at least 1,00,000 numbers between different schemes.
This ensures that application numbers do not overlap and remain uniquely identifiable across all schemes."""
        }),
        ('Important Dates', {
            'fields': (
                'application_open_date',
                'application_close_date',
                'successful_applicants_publish_date',
                'appeal_end_date',
                'lottery_result_date',
                'close_date'
            ),
            'description': 'Timeline for the scheme lifecycle'
        }),
        ('System Information', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )

    # def get_application_count(self, obj):
    #     """Calculate no of applications submitted so far"""
    #     return self.applications.count()
    
    def get_status(self, obj):
        """Calculate current status based on dates"""
        from django.utils import timezone
        now = timezone.now()
        
        if obj.close_date and now > obj.close_date:
            return "Closed"
        elif obj.lottery_result_date and now > obj.lottery_result_date:
            return "Lottery Announced"
        elif obj.appeal_end_date and now > obj.appeal_end_date:
            return "Lottery Pending"
        elif obj.successful_applicants_publish_date and now > obj.successful_applicants_publish_date:
            return "Appeal Period"
        elif obj.application_close_date and now > obj.application_close_date:
            return "Applications Under Review"
        elif obj.application_open_date and now >= obj.application_open_date:
            return "Applications Open"
        else:
            return "Coming Soon"
    
    get_status.short_description = 'Current Status'
    
    # Add date hierarchy for easy filtering
    date_hierarchy = 'application_open_date'




from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
import csv
from datetime import datetime
from .models import Application


class ApplicationAdmin(admin.ModelAdmin):
    # List display
    list_display = [
        'application_number',
        'applicant_name',
        'mobile_number',
        'scheme',
        'plot_category',
        'total_payable_amount',
        'payment_status_badge',
        'application_status_badge',
        'lottery_status_badge',
        'application_submission_date',
        'application_status',
        'payment_status',
        'lottery_status',
    ]
    
    # List filters
    list_filter = [
        'application_status',
        'payment_status',
        'lottery_status',
        'plot_category',
        'payment_mode',
        'id_type',
        'annual_income',
        'application_submission_date',
        'scheme',
    ]
    
    # Search fields
    search_fields = [
        'applicant_name',
        'mobile_number',
        'email',
        'id_number',
        'pan_number',
        'father_or_husband_name',
        'application_number',
    ]
    
    # Inline editing (list_editable)
    list_editable = [
        'payment_status',
        'application_status',
        'lottery_status',
    ]
    
    # Fieldsets for organized form view
    fieldsets = (
        ('Scheme Information', {
            'fields': ('scheme',)
        }),
        ('Basic Details', {
            'fields': ('applicant_name', 'father_or_husband_name', 'dob', 'mobile_number', 'email')
        }),
        ('Identity Details', {
            'fields': ('id_type', 'id_number', 'pan_number')
        }),
        ('Address Details', {
            'fields': (
                'permanent_address', 
                'permanent_address_pincode',
                'postal_address',
                'postal_address_pincode'
            )
        }),
        ('Income & Plot Category', {
            'fields': ('annual_income', 'plot_category')
        }),
        ('Fee Details', {
            'fields': ('registration_fees', 'processing_fees', 'total_payable_amount'),
            'classes': ('collapse',)
        }),
        ('Payment Details', {
            'fields': (
                'payment_mode',
                'dd_id_or_transaction_id',
                'dd_date_or_transaction_date',
                'dd_amount',
                'payee_account_holder_name',
                'payee_bank_name',
                'payment_proof',
            )
        }),
        ('Payment Status', {
            'fields': ('payment_status',),
            'classes': ('wide',)
        }),
        ('Refund Details', {
            'fields': (
                'refund_account_holder_name',
                'refund_account_number',
                'refund_bank_name',
                'refund_bank_ifsc',
                'refund_bank_branch_address',
                
            ),
            'classes': ('collapse',)
        }),
        ('Application Status', {
            'fields': ('application_status', 'rejection_remark', 'lottery_status'),
            'classes': ('wide',)
        }),
        ('Documents', {
            'fields': ('application_pdf',),
            'classes': ('collapse',)
        }),
    )
    
    # Read-only fields
    readonly_fields = [
        'plot_category',
        'registration_fees',
        'processing_fees',
        'total_payable_amount',
        'application_submission_date',
        'created_at',
        'updated_at',
    ]
    
    # Ordering
    ordering = ['-application_submission_date']
    
    # Items per page
    list_per_page = 50
    
    # Date hierarchy
    date_hierarchy = 'application_submission_date'
    
    # Actions
    actions = [
        'export_as_csv',
        'export_as_excel',
        'mark_payment_verified',
        'mark_application_accepted',
        'mark_application_rejected',
    ]
    
    # Custom colored badges for status fields
    def payment_status_badge(self, obj):
        colors = {
            'PENDING': 'orange',
            'VERIFIED': 'green',
            'FAILED': 'red',
        }
        color = colors.get(obj.payment_status, 'gray')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 3px;">{}</span>',
            color,
            obj.get_payment_status_display()
        )
    payment_status_badge.short_description = 'Payment Status'
    
    def application_status_badge(self, obj):
        colors = {
            'PENDING': 'orange',
            'ACCEPTED': 'green',
            'REJECTED': 'red',
        }
        color = colors.get(obj.application_status, 'gray')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 3px;">{}</span>',
            color,
            obj.get_application_status_display()
        )
    application_status_badge.short_description = 'Application Status'
    
    def lottery_status_badge(self, obj):
        colors = {
            'NOT_CONDUCTED': 'gray',
            'SELECTED': 'green',
            'NOT_SELECTED': 'red',
            'WAITLISTED': 'orange',
        }
        color = colors.get(obj.lottery_status, 'gray')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 3px;">{}</span>',
            color,
            obj.get_lottery_status_display()
        )
    lottery_status_badge.short_description = 'Lottery Status'
    
    # Export as CSV
    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        writer = csv.writer(response)
        writer.writerow(field_names)
        
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        
        return response
    export_as_csv.short_description = "Export selected as CSV"
    
    # Export as Excel (requires openpyxl: pip install openpyxl)
    def export_as_excel(self, request, queryset):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self.message_user(request, "Please install openpyxl: pip install openpyxl", level='error')
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        
        # Get field names
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        verbose_names = [field.verbose_name.title() for field in meta.fields]
        
        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, column_title in enumerate(verbose_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(field_names, 1):
                value = getattr(obj, field)
                # Convert datetime objects to strings
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
    export_as_excel.short_description = "Export selected as Excel"
    
    # Bulk action: Mark payment as verified
    def mark_payment_verified(self, request, queryset):
        updated = queryset.update(payment_status='VERIFIED')
        self.message_user(request, f'{updated} application(s) marked as payment verified.')
    mark_payment_verified.short_description = "Mark payment as verified"
    
    # Bulk action: Mark application as accepted
    def mark_application_accepted(self, request, queryset):
        updated = queryset.update(application_status='ACCEPTED')
        self.message_user(request, f'{updated} application(s) marked as accepted.')
    mark_application_accepted.short_description = "Mark application as accepted"
    
    # Bulk action: Mark application as rejected
    def mark_application_rejected(self, request, queryset):
        updated = queryset.update(application_status='REJECTED')
        self.message_user(request, f'{updated} application(s) marked as rejected.')
    mark_application_rejected.short_description = "Mark application as rejected"
    
    # Custom save behavior
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        # Add any custom save logic here if needed


# Register the model with the admin site
admin.site.register(Application, ApplicationAdmin)







